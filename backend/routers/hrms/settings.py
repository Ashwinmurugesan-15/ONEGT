from fastapi import APIRouter, HTTPException, Depends
from typing import List, Optional
import logging
import os
from services.google_sheets import sheets_service
from models.settings import (
    CompanySettings, Role, Permission, 
    LeaveGroup, LeaveType, LeavePolicy, Entitlement, Holiday
)
from config import settings
from middleware.auth_middleware import get_current_user, require_admin
from services.google_drive import drive_service
from fastapi import UploadFile, File
import uuid

logger = logging.getLogger("chrms.settings")

router = APIRouter()

def _gen_holiday_id():
    return f"HOL-{uuid.uuid4().hex[:6].upper()}"

@router.get("/company", response_model=CompanySettings)
async def get_company_settings(current_user = Depends(get_current_user)):
    """Return the single‑row company settings."""
    try:
        rows = sheets_service.get_all_records(settings.SETTINGS_COMPANY_SHEET)
        if rows:
            return CompanySettings(**rows[0])
        return CompanySettings()
    except Exception:
        return CompanySettings()

@router.post("/company", response_model=dict)
async def save_company_settings(payload: CompanySettings, current_user = Depends(require_admin)):
    """Upsert company settings — always a single row."""
    try:
        headers = list(payload.model_dump().keys())
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_COMPANY_SHEET, headers)
        existing = sheets_service.get_all_records(settings.SETTINGS_COMPANY_SHEET)
        row_data = list(payload.model_dump().values())
        if existing:
            sheets_service.update_row(settings.SETTINGS_COMPANY_SHEET, 2, row_data)
        else:
            sheets_service.append_row(settings.SETTINGS_COMPANY_SHEET, row_data)
        return {"success": True, "message": "Settings saved"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/upload-logo", response_model=dict)
async def upload_company_logo(file: UploadFile = File(...), current_user = Depends(require_admin)):
    """Upload company logo to Google Drive and store file ID."""
    try:
        if not file.content_type or not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="Only image files are allowed")
        
        contents = await file.read()
        if len(contents) > 5 * 1024 * 1024:  # 5MB limit
            raise HTTPException(status_code=400, detail="File size must be under 5MB")
        
        folder_id = settings.ROOT_FOLDER_ID or os.getenv("DRIVE_PEOPLES_FOLDER_ID") or None
        ext = file.filename.split(".")[-1] if file.filename else "png"
        filename = f"company_logo.{ext}"
        
        file_id = drive_service.upload_file_binary(contents, filename, file.content_type, folder_id)
        if not file_id:
            raise HTTPException(status_code=500, detail="Failed to upload to Google Drive")
        
        # Update company settings with the new logo file ID
        try:
            records = sheets_service.get_all_records(settings.SETTINGS_COMPANY_SHEET)
            if records:
                row = records[0]
                values = [
                    row.get("name", ""), file_id,
                    row.get("primary_color", "#3b82f6"), row.get("secondary_color", "#10b981"),
                    row.get("address", ""), row.get("website", ""),
                    row.get("email", ""), row.get("phone", "")
                ]
                sheets_service.update_row(settings.SETTINGS_COMPANY_SHEET, 2, values)
        except Exception as ex:
            logger.warning(f"Logo uploaded but failed to update settings sheet: {ex}")
        
        return {"success": True, "file_id": file_id, "message": "Logo uploaded successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading logo: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/roles", response_model=List[Role])
async def get_roles(current_user = Depends(get_current_user)):
    try:
        records = sheets_service.get_all_records(settings.SETTINGS_ROLES_SHEET)
        roles = []
        for r in records:
            # Ensure is_default has a fallback for older sheets without the column
            if "is_default" not in r:
                r["is_default"] = False
            elif isinstance(r["is_default"], str):
                r["is_default"] = r["is_default"].upper() == "TRUE"
            roles.append(Role(**r))
        return roles
    except Exception as e:
        import logging
        logging.getLogger("chrms").error(f"Error loading roles: {e}")
        return []

@router.post("/roles", response_model=dict)
async def create_role(role: dict, current_user = Depends(require_admin)):
    try:
        role_name = role.get("name", "")
        role_desc = role.get("description", "")
        is_default = role.get("is_default", False)
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_ROLES_SHEET, ["id", "name", "description", "is_default"])
        # Generate GTRL sequence ID
        existing = sheets_service.get_all_records(settings.SETTINGS_ROLES_SHEET)
        max_seq = 0
        for r in existing:
            rid = str(r.get("id", ""))
            if rid.startswith("GTRL") and rid[4:].isdigit():
                max_seq = max(max_seq, int(rid[4:]))
        role_id = f"GTRL{max_seq + 1:04d}"
        sheets_service.append_row(settings.SETTINGS_ROLES_SHEET, [role_id, role_name, role_desc, is_default])
        return {"success": True, "message": "Role created", "id": role_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/roles/{role_id}/set-default", response_model=dict)
async def set_default_role(role_id: str, current_user = Depends(require_admin)):
    try:
        records = sheets_service.get_all_records(settings.SETTINGS_ROLES_SHEET)
        updated_rows = []
        for r in records:
            r["is_default"] = True if str(r.get("id", "")) == role_id else False
            updated_rows.append([r.get("id", ""), r.get("name", ""), r.get("description", ""), r["is_default"]])
        sheets_service.clear_sheet(settings.SETTINGS_ROLES_SHEET)
        all_values = [["id", "name", "description", "is_default"]] + updated_rows
        sheets_service.update_values(settings.SETTINGS_ROLES_SHEET, all_values)
        return {"success": True, "message": f"Role {role_id} set as default"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/roles/seed-admin", response_model=dict)
async def seed_admin_role(current_user = Depends(require_admin)):
    """Create an Admin role with full read/write/all permissions on every module and page."""
    try:
        # All modules and pages
        MODULES = {
            "HRMS": ["Dashboard", "Associates", "Payroll", "Asset Management", "Org Chart",
                     "Projects", "Allocations", "Timesheets", "Expenses", "Currency Rates",
                     "Personal Info", "Pay Structure"],
            "CRMS": ["Dashboard", "Leads", "Opportunities", "Deals", "Customers", "Contacts",
                     "Invoices", "Finance View", "Tasks", "Call Logs"],
            "TalentManagement": ["Dashboard", "Job Postings", "Candidates", "Interviews",
                                 "Training Programs", "Performance Reviews", "Goals & OKRs",
                                 "Succession Planning", "Skill Matrix"],
            "AssessmentPortal": ["Dashboard", "All Assessments", "Create Assessment", "Question Bank",
                                 "Candidates", "Invitations", "Reports", "Analytics"],
            "PMS": ["Dashboard", "Goal Templates", "Appraisal Cycles", "My Appraisals"]
        }

        # Step 1: Create Admin role if not exists
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_ROLES_SHEET, ["id", "name", "description", "is_default"])
        existing_roles = sheets_service.get_all_records(settings.SETTINGS_ROLES_SHEET)
        admin_role = next((r for r in existing_roles if str(r.get("name", "")).lower() == "admin"), None)

        if admin_role:
            admin_role_id = admin_role["id"]
        else:
            max_seq = 0
            for r in existing_roles:
                rid = str(r.get("id", ""))
                if rid.startswith("GTRL") and rid[4:].isdigit():
                    max_seq = max(max_seq, int(rid[4:]))
            admin_role_id = f"GTRL{max_seq + 1:04d}"
            sheets_service.append_row(settings.SETTINGS_ROLES_SHEET, [admin_role_id, "Admin", "Full access to all modules and pages", False])

        # Step 2: Build admin permissions
        admin_perms = []
        for cap_id, pages in MODULES.items():
            for page in pages:
                admin_perms.append([admin_role_id, cap_id, page, True, True, "all"])

        # Step 3: Merge with existing non-admin permissions
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_PERMISSIONS_SHEET, [
            "role_id", "capability_id", "page_id", "can_read", "can_write", "scope"
        ])
        existing_perms = sheets_service.get_all_records(settings.SETTINGS_PERMISSIONS_SHEET)
        other_perm_rows = [
            [p.get("role_id", ""), p.get("capability_id", ""), p.get("page_id", ""),
             p.get("can_read", False), p.get("can_write", False), p.get("scope", "associate")]
            for p in existing_perms if p.get("role_id") != admin_role_id
        ]

        # Step 4: Write all permissions
        header = ["role_id", "capability_id", "page_id", "can_read", "can_write", "scope"]
        sheets_service.clear_sheet(settings.SETTINGS_PERMISSIONS_SHEET)
        sheets_service.update_values(settings.SETTINGS_PERMISSIONS_SHEET, [header] + other_perm_rows + admin_perms)

        total_pages = sum(len(pages) for pages in MODULES.values())
        return {"success": True, "message": f"Admin role '{admin_role_id}' created with {total_pages} full-access permissions", "id": admin_role_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/leave-groups", response_model=List[LeaveGroup])
async def get_leave_groups(current_user = Depends(get_current_user)):
    try:
        records = sheets_service.get_all_records(settings.SETTINGS_LEAVE_GROUPS_SHEET)
        groups = []
        for r in records:
            # Normalize keys: lowercase + spaces → underscores
            norm = {k.strip().lower().replace(' ', '_'): v for k, v in r.items()}
            # Handle old-format records that have 'id' instead of 'code'
            if 'code' not in norm and 'id' in norm:
                norm['code'] = norm.pop('id')
            if not norm.get('attachment_mandatory'):
                norm['attachment_mandatory'] = 'No'
            groups.append(LeaveGroup(**norm))
        return groups
    except Exception as e:
        logger.error(f"Error loading leave groups: {e}")
        return []

@router.post("/leave-groups", response_model=dict)
async def create_leave_group(group: LeaveGroup, current_user = Depends(require_admin)):
    try:
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_LEAVE_GROUPS_SHEET, ["Code", "Name", "Description", "Attachment Mandatory"])
        sheets_service.append_row(settings.SETTINGS_LEAVE_GROUPS_SHEET, [group.code, group.name, group.description or "", group.attachment_mandatory])
        return {"success": True, "message": "Leave Group created", "code": group.code}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/leave-groups/{group_code}", response_model=dict)
async def update_leave_group(group_code: str, group: LeaveGroup, current_user = Depends(require_admin)):
    """Update an existing leave group."""
    try:
        # Try both capitalized and lowercase header names
        row_idx = sheets_service.find_row_index(settings.SETTINGS_LEAVE_GROUPS_SHEET, "Code", group_code)
        if not row_idx:
            row_idx = sheets_service.find_row_index(settings.SETTINGS_LEAVE_GROUPS_SHEET, "code", group_code)
        if not row_idx:
            raise HTTPException(status_code=404, detail="Leave group not found")
        sheets_service.update_row(settings.SETTINGS_LEAVE_GROUPS_SHEET, row_idx, [
            group_code, group.name, group.description or "", group.attachment_mandatory
        ])
        return {"success": True, "message": "Leave Group updated"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/leave-types", response_model=List[LeaveType])
async def get_leave_types(current_user = Depends(get_current_user)):
    try:
        records = sheets_service.get_all_records(settings.SETTINGS_LEAVE_TYPES_SHEET)
        types = []
        for r in records:
            norm = {k.strip().lower().replace(' ', '_'): v for k, v in r.items()}
            if not norm.get('code'):
                norm['code'] = norm.get('id', '')
            # Remove old 'id' key if present since model no longer has it
            norm.pop('id', None)
            if not norm.get('is_leave'):
                norm['is_leave'] = 'Yes'
            if not norm.get('pay_type'):
                norm['pay_type'] = 'PAID'
            if not norm.get('active'):
                norm['active'] = 'Yes'
            if not norm.get('is_adjustable'):
                norm['is_adjustable'] = 'No'
            types.append(LeaveType(**norm))
        return types
    except Exception as e:
        logger.error(f"Error loading leave types: {e}")
        return []

@router.post("/leave-types", response_model=dict)
async def create_leave_type(lt: LeaveType, current_user = Depends(require_admin)):
    try:
        headers = ["code", "name", "group_id", "is_leave", "pay_type", "active", "is_adjustable", "description"]
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_LEAVE_TYPES_SHEET, headers)
        sheets_service.append_row(settings.SETTINGS_LEAVE_TYPES_SHEET, [
            lt.code, lt.name, lt.group_id, lt.is_leave, lt.pay_type,
            lt.active, lt.is_adjustable, lt.description or ""
        ])
        return {"success": True, "message": "Leave Type created", "code": lt.code}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/leave-types/{type_code}", response_model=dict)
async def update_leave_type(type_code: str, lt: LeaveType, current_user = Depends(require_admin)):
    """Update an existing leave type."""
    try:
        row_idx = sheets_service.find_row_index(settings.SETTINGS_LEAVE_TYPES_SHEET, "code", type_code)
        if not row_idx:
            raise HTTPException(status_code=404, detail="Leave type not found")
        sheets_service.update_row(settings.SETTINGS_LEAVE_TYPES_SHEET, row_idx, [
            type_code, lt.name, lt.group_id, lt.is_leave, lt.pay_type,
            lt.active, lt.is_adjustable, lt.description or ""
        ])
        return {"success": True, "message": "Leave Type updated"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/permissions", response_model=List[Permission])
async def get_permissions(current_user = Depends(get_current_user)):
    try:
        records = sheets_service.get_all_records(settings.SETTINGS_PERMISSIONS_SHEET)
        return [Permission(**r) for r in records]
    except Exception:
        return []

@router.post("/permissions", response_model=dict)
async def update_permissions(perms: List[Permission], current_user = Depends(require_admin)):
    try:
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_PERMISSIONS_SHEET, [
            "role_id", "capability_id", "page_id", "can_read", "can_write", "scope"
        ])
        # Clear existing data and rewrite all permissions (including scope)
        sheets_service.clear_sheet(settings.SETTINGS_PERMISSIONS_SHEET)
        header = ["role_id", "capability_id", "page_id", "can_read", "can_write", "scope"]
        rows = [[p.role_id, p.capability_id, p.page_id, p.can_read, p.can_write, p.scope] for p in perms]
        sheets_service.update_values(settings.SETTINGS_PERMISSIONS_SHEET, [header] + rows)
        return {"success": True, "message": "Permissions updated"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/leave-policies", response_model=List[LeavePolicy])
async def get_leave_policies(current_user = Depends(get_current_user)):
    try:
        records = sheets_service.get_all_records(settings.SETTINGS_LEAVE_POLICIES_SHEET)
        return [LeavePolicy(**r) for r in records]
    except Exception:
        return []

@router.post("/leave-policies", response_model=dict)
async def update_leave_policy(policy: LeavePolicy, current_user = Depends(require_admin)):
    try:
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_LEAVE_POLICIES_SHEET, [
            "leave_type_id", "calculation_basis", "from_date", "to_date", 
            "prorate_based_on_join_date", "calculate_earned_leave", "expiry_months", 
            "avail_earned_leave_by", "leave_limit_option", "entitlement_calc_basis", 
            "allow_viewer_calendar", "permission_day_wise", "permission_days_per_month", 
            "permission_hours_per_day", "permission_hour_wise", "permission_hours_per_month"
        ])
        values = [
            policy.leave_type_id, policy.calculation_basis, policy.from_date, policy.to_date,
            policy.prorate_based_on_join_date, policy.calculate_earned_leave, policy.expiry_months,
            policy.avail_earned_leave_by, policy.leave_limit_option, policy.entitlement_calc_basis,
            policy.allow_viewer_calendar, policy.permission_day_wise, policy.permission_days_per_month,
            policy.permission_hours_per_day, policy.permission_hour_wise, policy.permission_hours_per_month
        ]
        row_idx = sheets_service.find_row_index(settings.SETTINGS_LEAVE_POLICIES_SHEET, "leave_type_id", policy.leave_type_id)
        if row_idx:
            sheets_service.update_row(settings.SETTINGS_LEAVE_POLICIES_SHEET, row_idx, values)
        else:
            sheets_service.append_row(settings.SETTINGS_LEAVE_POLICIES_SHEET, values)
        return {"success": True, "message": "Leave policy updated"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Leave Entitlements ────────────────────────────────────────

ENT_HEADERS = ["group_code", "entitlement_type", "from_year", "to_year", "entitlement_days"]

@router.get("/leave-entitlements", response_model=List[Entitlement])
async def get_leave_entitlements(current_user = Depends(get_current_user)):
    try:
        records = sheets_service.get_all_records(settings.SETTINGS_LEAVE_ENTITLEMENTS_SHEET)
        entitlements = []
        for r in records:
            entitlements.append(Entitlement(
                group_code=str(r.get("group_code", "")),
                entitlement_type=str(r.get("entitlement_type", "common")),
                from_year=int(r.get("from_year", 0)),
                to_year=int(r.get("to_year", 0)),
                entitlement_days=float(r.get("entitlement_days", 0))
            ))
        return entitlements
    except Exception:
        return []

from pydantic import BaseModel as PydanticBaseModel

class SaveEntitlementsRequest(PydanticBaseModel):
    group_code: str
    entitlement_type: str = "common"
    rows: List[Entitlement]

@router.post("/leave-entitlements", response_model=dict)
async def save_leave_entitlements(req: SaveEntitlementsRequest, current_user = Depends(require_admin)):
    try:
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_LEAVE_ENTITLEMENTS_SHEET, ENT_HEADERS)
        # Delete existing rows for this group_code (reverse order)
        try:
            records = sheets_service.get_all_records(settings.SETTINGS_LEAVE_ENTITLEMENTS_SHEET)
            rows_to_delete = []
            for idx, r in enumerate(records):
                if str(r.get("group_code", "")) == req.group_code:
                    rows_to_delete.append(idx + 2)
            for row_num in reversed(rows_to_delete):
                sheets_service.delete_row(settings.SETTINGS_LEAVE_ENTITLEMENTS_SHEET, row_num)
        except Exception:
            pass
        # Append new rows
        for ent in req.rows:
            sheets_service.append_row(settings.SETTINGS_LEAVE_ENTITLEMENTS_SHEET, [
                req.group_code, req.entitlement_type, ent.from_year, ent.to_year, ent.entitlement_days
            ])
        sheets_service.clear_cache(settings.SETTINGS_LEAVE_ENTITLEMENTS_SHEET)
        return {"success": True, "message": "Entitlements saved"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/leave-entitlements/{group_code}", response_model=dict)
async def delete_leave_entitlements(group_code: str, current_user = Depends(require_admin)):
    try:
        records = sheets_service.get_all_records(settings.SETTINGS_LEAVE_ENTITLEMENTS_SHEET)
        rows_to_delete = []
        for idx, r in enumerate(records):
            if str(r.get("group_code", "")) == group_code:
                rows_to_delete.append(idx + 2)
        for row_num in reversed(rows_to_delete):
            sheets_service.delete_row(settings.SETTINGS_LEAVE_ENTITLEMENTS_SHEET, row_num)
        sheets_service.clear_cache(settings.SETTINGS_LEAVE_ENTITLEMENTS_SHEET)
        return {"success": True, "message": "Entitlements deleted"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Holidays ──────────────────────────────────────────────────

HOLIDAY_HEADERS = [
    "id", "name", "date", "day", "holiday_type",
    "applicable_to", "year", "description"
]

@router.get("/holidays")
async def get_holidays(year: int = None, current_user = Depends(get_current_user)):
    """Get all configured holidays, optionally filtered by year."""
    try:
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_HOLIDAYS_SHEET, HOLIDAY_HEADERS)
        records = sheets_service.get_all_records(settings.SETTINGS_HOLIDAYS_SHEET)
        if year:
            records = [r for r in records if str(r.get("year")) == str(year)]
        # Sort by date
        records.sort(key=lambda r: r.get("date", ""))
        return records
    except Exception as e:
        logger.error(f"Error fetching holidays: {e}")
        return []

from fastapi import UploadFile, File

@router.post("/holidays/upload", response_model=dict)
async def upload_holidays(file: UploadFile = File(...), current_user = Depends(require_admin)):
    """Bulk upload holidays from CSV or XLSX file.
    Expected columns: name, date, holiday_type, applicable_to, description
    """
    import io
    from datetime import datetime as dt

    try:
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_HOLIDAYS_SHEET, HOLIDAY_HEADERS)
        contents = await file.read()
        rows = []

        fname = (file.filename or "").lower()
        if fname.endswith(".csv"):
            import csv
            text = contents.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append(row)
        elif fname.endswith(".xlsx") or fname.endswith(".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            ws = wb.active
            headers_row = [str(c.value or "").strip().lower() for c in ws[1]]
            for r in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, val in enumerate(r):
                    if i < len(headers_row):
                        row_dict[headers_row[i]] = val
                if any(v for v in row_dict.values() if v):
                    rows.append(row_dict)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Use .csv or .xlsx")

        added = 0
        for row in rows:
            name = str(row.get("name", "") or "").strip()
            date_val = row.get("date", "")
            if not name or not date_val:
                continue

            # Parse date
            date_str = ""
            day_name = ""
            year_val = 0
            if isinstance(date_val, str):
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
                    try:
                        d = dt.strptime(date_val.strip(), fmt)
                        date_str = d.strftime("%Y-%m-%d")
                        day_name = d.strftime("%A")
                        year_val = d.year
                        break
                    except ValueError:
                        continue
            else:
                try:
                    d = date_val if isinstance(date_val, dt) else dt.combine(date_val, dt.min.time())
                    date_str = d.strftime("%Y-%m-%d")
                    day_name = d.strftime("%A")
                    year_val = d.year
                except Exception:
                    continue

            if not date_str:
                continue

            hid = _gen_holiday_id()
            h_type = str(row.get("holiday_type", "") or row.get("type", "") or "Company").strip() or "Company"
            applicable = str(row.get("applicable_to", "") or "All").strip() or "All"
            desc = str(row.get("description", "") or "").strip()

            sheets_service.append_row(settings.SETTINGS_HOLIDAYS_SHEET, [
                hid, name, date_str, day_name, h_type, applicable, year_val, desc
            ])
            added += 1

        sheets_service.clear_cache(settings.SETTINGS_HOLIDAYS_SHEET)
        return {"success": True, "message": f"{added} holidays imported successfully", "count": added}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading holidays: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/holidays", response_model=dict)
async def create_holiday(holiday: Holiday, current_user = Depends(require_admin)):
    """Add a new holiday."""
    try:
        sheets_service.create_sheet_if_not_exists(settings.SETTINGS_HOLIDAYS_SHEET, HOLIDAY_HEADERS)
        hid = holiday.id if holiday.id else _gen_holiday_id()
        # Compute day of week from date
        from datetime import datetime as dt
        day_name = ""
        try:
            d = dt.strptime(holiday.date, "%Y-%m-%d")
            day_name = d.strftime("%A")
            if not holiday.year:
                holiday.year = d.year
        except Exception:
            day_name = holiday.day

        sheets_service.append_row(settings.SETTINGS_HOLIDAYS_SHEET, [
            hid, holiday.name, holiday.date, day_name,
            holiday.holiday_type, holiday.applicable_to,
            holiday.year, holiday.description or ""
        ])
        return {"success": True, "id": hid, "message": "Holiday added successfully"}
    except Exception as e:
        logger.error(f"Error creating holiday: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/holidays/{holiday_id}", response_model=dict)
async def update_holiday(holiday_id: str, holiday: Holiday, current_user = Depends(require_admin)):
    """Update an existing holiday."""
    try:
        row_idx = sheets_service.find_row_index(settings.SETTINGS_HOLIDAYS_SHEET, "id", holiday_id)
        if not row_idx:
            raise HTTPException(status_code=404, detail="Holiday not found")

        from datetime import datetime as dt
        day_name = ""
        try:
            d = dt.strptime(holiday.date, "%Y-%m-%d")
            day_name = d.strftime("%A")
        except Exception:
            day_name = holiday.day

        sheets_service.update_row(settings.SETTINGS_HOLIDAYS_SHEET, row_idx, [
            holiday_id, holiday.name, holiday.date, day_name,
            holiday.holiday_type, holiday.applicable_to,
            holiday.year, holiday.description or ""
        ])
        return {"success": True, "message": "Holiday updated"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating holiday: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/holidays/{holiday_id}", response_model=dict)
async def delete_holiday(holiday_id: str, current_user = Depends(require_admin)):
    """Delete a holiday."""
    try:
        row_idx = sheets_service.find_row_index(settings.SETTINGS_HOLIDAYS_SHEET, "id", holiday_id)
        if not row_idx:
            raise HTTPException(status_code=404, detail="Holiday not found")
        sheets_service.delete_row(settings.SETTINGS_HOLIDAYS_SHEET, row_idx)
        return {"success": True, "message": "Holiday deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting holiday: {e}")
        raise HTTPException(status_code=500, detail=str(e))

